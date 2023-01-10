from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import itertools
import sys
import os
from datetime import date
from openpyxl.styles import colors
#import automate as auto
import time
import shutil
import subprocess
# TO DO to finish the automated daily postcard code sales report
# 1. Consider changing PT and PD to "Pay Today" and "Post Date"
# 	1.1 If have to change this change print associated and count codes to match the larger concate
#	1.2 Adjust map to add right values
#	1.3 Katie said we dont need to change
# 2. Figure out solution to access document to run this script
# 	2.1 API to access shared oneDrive file
# 	2.2 See if there is a power automate to download a copy of the spreadsheet to a shared drive (will require alex)
#   2.3 Ask Alex Dill if I can make a power automate on his account
#	2.4 Setting up macro to copy file to katie everyday
# 3. Find a solution to the limit of rows to read on the spreadsheets
# 	3.1. If going to put a large number as limit how will I differentiate postcard sales with no code entries from just the empty cells
#		3.1.2 EASY actually lets just reverse the order of for loops when loading data so that we can make a var to to turn on when sign of blank post card, but then confirm the break out of loop when you see blank terms
# 4. Cell color handling enabled and no longer saving postcards that have no refID

def printAssociated (Pcodes, terms, Pcodesfull, match):
	for (x,y,z) in zip(Pcodes, terms, Pcodesfull):
		if x == match[0:2] and y == match[3:5]:
			print(z)
	#print(" ")

def countCodes(headers, Pcodes, Pcount, terms):
	for i in headers:
		count = 0
		for (x,y) in zip(Pcodes,terms):
			if x == i[0:2] and y == i[3:5]:
				count = count + 1
				j = count
		#j = count
		Pcount.append(count)

def getData(worksheet, terms, Pcodes, Pcodesfull):
	ws = wb[worksheet]
	flop = ""
	for row in range (5, 160):
		for col in range (5,7):
		#tcount = 0
		#pcount = 0
			char = get_column_letter(col)
			if char == "E":
				
				if (ws[char + str(row)].value is None):
					break

				elif (ws[char + str(row)].value == "Pay Today"):
					flop = "PT"
					#terms.append("PT")

					#tcount += 1
				elif (ws[char + str(row)].value == "Post Date"):
					flop = "PD"
					#terms.append("PD")
					#tcount += 1
			else:
				#print(str(ws[char + str(row)].fill.fgColor.rgb))
				if (ws[char+ str(row)].fill.fgColor.rgb == "FFFF0000"):
					break
				else:
					#print(ws[char + str(row)].value)
					if (ws[char + str(row)].value) is None:
						break
						#Pcodes.append("00")
						#Pcodesfull.append("N0C0DE")
					else:
						if flop == "PT":
							terms.append("PT")
						elif flop == "PD":
							terms.append("PD")

						flop = ""
						val = type(ws[char+str(row)].value)
						#print(val)
						if isinstance(ws[char+str(row)].value, int):
							Pcodes.append("XX")
							Pcodesfull.append("XX" + str(ws[char + str(row)].value))
						elif (ws[char +str(row)].value[0:1] == " "):
							Pcodes.append(ws[char + str(row)].value[1:3])
							Pcodesfull.append(ws[char + str(row)].value[1:])
						else:
							Pcodes.append(ws[char + str(row)].value[0:2].upper())
							Pcodesfull.append(ws[char + str(row)].value[0:2].upper() + ws[char + str(row)].value[2:])

def manipData(Pcodes, Pcodesu, headers):
	for x in Pcodes:
		duplicate = 0
		for y in Pcodesu:
			if x == y:
				duplicate = 1

		if duplicate == 0:
			Pcodesu.append(x)

	Pcodesu.sort()
	Pcodesdup = Pcodesu
	#print (Pcodesdup)
	#print (Pcodes)
	#print (terms)
	result = map(lambda x: x + " PT", Pcodesu)
	result2 = map(lambda x: x + " PD", Pcodesdup)
	a = list(result)
	b = list(result2)
	headers[:] = a + b
	headers.sort()
	#print(result2)
	#print(result)
	#print(headers)
	countCodes(headers, Pcodes, Pcount, terms)



# def countPayment (terms):
# 	# global PTcount
# 	# global PDcount
# 	# PTcount = 0
# 	# PDcount = 0

# 	for l in terms:
# 		if l == "PT":
# 			PTcount += 1
# 		else:
# 			PDcount += 1

# 	print(PTcount)

#print(headers)
def returnText(worksheet, headers, Pcount, Pcodes, terms, Pcodesfull):
	PTcount = 0
	PDcount = 0
	
	for l in terms:
		if l == "PT":
			PTcount += 1
		else:
			PDcount += 1
	
	original_stdout = sys.stdout
	email = 'emails\\email_' + worksheet + "_" + str(year) +'.txt'
	with open(email, 'w') as f:
		sys.stdout = f
		#print(headers)
		for (x,y) in zip(headers, Pcount):
			#print("goofy")
			if (y == 0):
				continue
			else:
				print (x + ": " + str(y))
				#print(" ")
				printAssociated(Pcodes, terms, Pcodesfull, x)
		#print("after")
		print("")
		print("Total Paid Today: " + str(PTcount))
		print("Total Postdate: " + str(PDcount), end="")
		#print("")
		sys.stdout = original_stdout
	#f.close()


def printText(emailLoc):
	with open(emailLoc, 'r') as f:
		contents = f.read()
		#contents.strip('/r')
		#contents = contents.strip
		print(contents)

Pcodesfull = []
Pcodes = []
terms = []
used = []
Pcodesu = []
Pcount = []
Pcodesdup = []
headers = []
currsheet = "January 2023 Sales & QC Log.xlsx"

if (date.today().weekday() == 1):
	pass

else:
	today = date.today()
	today = today.replace(day=today.day -1 )
	date = today.strftime("%m%d")
	year = today.year
	date = "0107"
	#print(date)
	myfile="C:\\Users\\mbreden\\Desktop\\test-react-app\\Backend\\data\\" + currsheet
	myfile2="C:\\Users\\mbreden\\Downloads\\" + currsheet
	if os.path.isfile(myfile):
	 	os.remove(myfile)
	 	#print("Data File removed")
	if os.path.isfile(myfile2):
	 	os.remove(myfile2)
	 	#print("Downloads File removed")

	#auto.do_this()
	working_dir = "C:\\Users\\mbreden\\Desktop\\test-react-app\\Backend\\scripts"
	#print("Starting Download...")
	subprocess.check_call(['node', 'sales.js'], cwd=working_dir)
	time.sleep(20)
	#print("Finished Download...")
	os.rename(myfile2, myfile)
	os.chdir("C:\\Users\\mbreden\\Desktop\\test-react-app\\Backend\\data")
	wb = load_workbook(currsheet)
	#date = "0102"


	#print(type(date))
	#print(terms)
	#print(Pcodesu)
	#date = "1231"  #Use this when trying to run report for specific date
	getData(date, terms, Pcodes, Pcodesfull)
	manipData(Pcodes, Pcodesu, headers)
#countPayment(terms)
	returnText(date, headers, Pcount, Pcodes, terms, Pcodesfull)
	PTcount = 0
	PDcount = 0
	
	for l in terms:
		if l == "PT":
			PTcount += 1
		else:
			PDcount += 1
	#print(PTcount)
	#print(PDcount)
	#print(Pcount)
	#print(headers)
	email = 'email_' + date + "_"+ str(year) +  '.txt'
	emailLoc = 'C:\\Users\\mbreden\\Desktop\\test-react-app\\Backend\\data\\emails\\' + email

	printText(emailLoc)

	#shutil.copy(emailLoc, "\\\\DSI-WD-kellis-p\\c$\\Users\\kellis.DSI\\Desktop\\PostcodeSummary")
	#shutil.copy(emailLoc, "C:\\Users\\mbreden\\Desktop\\IT Shared")




#print(ws['C15'].value)

#rowcount froggy try to make big and ask if null to breaksed
#def getData(worksheet, terms, Pcodes, Pcodesfull):
#	ws = wb[worksheet]
#	for row in range (5, 100):
#		for col in range (5,7):
#		#tcount = 0
#		#pcount = 0
#			char = get_column_letter(col)
#			if char == "E":
#				if (ws[char + str(row)].value is None):
#					break
#
#				elif (ws[char + str(row)].value == "Pay Today"):
#					terms.append("PT")
#					#tcount += 1
#				else:
#					terms.append("PD")
#					#tcount += 1
#			else:
#				print(str(ws[char + str(row)].fill.fgColor.rgb))
#				if (ws[char+ str(row)].fill.fgColor.rgb == "FFFF0000"):
#					break
#				else:
#					if (ws[char + str(row)].value) is None:
#						break
						#Pcodes.append("00")
						#Pcodesfull.append("N0C0DE")
#					else:
#						if (ws[char +str(row)].value[0:1] == " "):
#							Pcodes.append(ws[char + str(row)].value[1:3])
#							Pcodesfull.append(ws[char + str(row)].value)
#						else:
#							Pcodes.append(ws[char + str(row)].value[0:2])
#							Pcodesfull.append(ws[char + str(row)].value)


#print (Pcodes)
#print (terms)
#print (date)
#def manipData(Pcodes, Pcodesu, headers):
#	for x in Pcodes:
#		duplicate = 0
#		for y in Pcodesu:
#			if x == y:
#				duplicate = 1
#
#		if duplicate == 0:
#			Pcodesu.append(x)

#	Pcodesu.sort()
#	Pcodesdup = Pcodesu
#	#print (Pcodesdup)
#	print (Pcodes)
#	print (terms)
#	result = map(lambda x: x + " PT", Pcodesu)
#	result2 = map(lambda x: x + " PD", Pcodesdup)
#	a = list(result)
#	b = list(result2)
#	headers = a + b
#	headers.sort()
	#print(result2)
	#print(result)
#	print(headers)
#	countCodes(headers, Pcodes, Pcount, terms)
#print(terms)

#print(Pcount)
#def countPayment (terms):
#	PTcount = 0
#	PDcount = 0
#	for l in terms:
#		if l == "PT":
#			PTcount += 1
#		else:
#			PDcount += 1

#def returnText(worksheet, headers, Pcount, Pcodes, terms, Pcodesfull):
#	original_stdout = sys.stdout
#	with open('email_' + worksheet + '.txt', 'w') as f:
#		sys.stdout = f
#		for (x,y) in zip(headers, Pcount):
#			print (x + ": " + str(y))
#			print(" ")
#			printAssociated(Pcodes, terms, Pcodesfull, x)
#		print(" ")
#		print("Total Paid Today: " + str(PTcount))
#		print("Total Postdate: " + str(PDcount))
#		sys.stdout = original_stdout


#final.sort()
#print(final)
#wb.save('April_Sales.xlsx')



